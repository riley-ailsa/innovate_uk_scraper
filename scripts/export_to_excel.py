#!/usr/bin/env python3
"""
Scrape Innovate UK competitions and export to Excel.

Usage:
    python scrape_to_excel.py [--limit N] [--output FILENAME]

Examples:
    python scrape_to_excel.py                    # Scrape with defaults
    python scrape_to_excel.py --limit 5          # Scrape 5 competitions
    python scrape_to_excel.py --output test.xlsx # Custom output file
"""

import sys
import argparse
import logging
from pathlib import Path
from datetime import datetime

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from src.ingest.innovateuk_competition import InnovateUKCompetitionScraper
from src.normalize.innovate_uk import normalize_scraped_competition

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def load_urls_from_file(filepath: str = "innovate_uk_urls.txt") -> list:
    """
    Load competition URLs from a text file.

    Args:
        filepath: Path to the URLs file

    Returns:
        List of URLs (ignores comments and blank lines)
    """
    urls = []
    file_path = Path(__file__).parent / filepath

    if not file_path.exists():
        logger.warning(f"URL file not found: {file_path}")
        return urls

    with open(file_path, 'r') as f:
        for line in f:
            line = line.strip()
            # Skip empty lines and comments
            if line and not line.startswith('#'):
                urls.append(line)

    return urls


def scrape_competitions(urls: list, limit: int = None) -> list:
    """
    Scrape competition data from Innovate UK.

    Args:
        urls: List of competition URLs to scrape
        limit: Maximum number to scrape (None = all)

    Returns:
        List of scraped competition data dicts
    """
    scraper = InnovateUKCompetitionScraper()
    results = []

    urls_to_process = urls[:limit] if limit else urls

    for i, url in enumerate(urls_to_process, 1):
        logger.info(f"[{i}/{len(urls_to_process)}] Scraping: {url}")

        try:
            scraped = scraper.scrape_competition(url)
            comp = scraped.competition

            # Normalize to get computed fields (competition_type, project_funding, etc.)
            grant, _ = normalize_scraped_competition(scraped, [])

            # Convert to dict for Excel export
            result = {
                'id': comp.id,
                'external_id': comp.external_id,
                'title': grant.title,  # Use cleaned title from grant
                'competition_type': grant.competition_type,
                'description': comp.description[:500] if comp.description else '',
                'url': comp.base_url,
                'opens_at': comp.opens_at.isoformat() if comp.opens_at else '',
                'closes_at': comp.closes_at.isoformat() if comp.closes_at else '',
                'total_fund': grant.total_fund or '',
                'project_size': comp.project_size or '',
                'project_funding_min': grant.project_funding_min,
                'project_funding_max': grant.project_funding_max,
                'expected_winners': grant.expected_winners,
                'funding_rules': str(comp.funding_rules) if comp.funding_rules else '',
                'sections_count': len(scraped.sections),
                'resources_count': len(scraped.resources),
            }

            results.append(result)
            logger.info(f"  ✓ {grant.title}")

        except Exception as e:
            logger.error(f"  ✗ Error: {e}")
            continue

    return results


def export_to_excel(data: list, filename: str):
    """
    Export scraped data to Excel file.

    Args:
        data: List of competition dicts
        filename: Output Excel filename
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Innovate UK Competitions"

    # Define columns
    columns = [
        ('ID', 10),
        ('External ID', 12),
        ('Title', 50),
        ('Type', 10),
        ('Description', 60),
        ('URL', 40),
        ('Opens', 20),
        ('Closes', 20),
        ('Total Fund', 25),
        ('Project Size', 25),
        ('Min Award', 15),
        ('Max Award', 15),
        ('Est. Winners', 12),
        ('Funding Rules', 30),
        ('Sections', 10),
        ('Resources', 10),
    ]

    # Header row styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Write headers
    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, record in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=record.get('id', ''))
        ws.cell(row=row_idx, column=2, value=record.get('external_id', ''))
        ws.cell(row=row_idx, column=3, value=record.get('title', ''))
        ws.cell(row=row_idx, column=4, value=record.get('competition_type', ''))
        ws.cell(row=row_idx, column=5, value=record.get('description', ''))
        ws.cell(row=row_idx, column=6, value=record.get('url', ''))
        ws.cell(row=row_idx, column=7, value=record.get('opens_at', ''))
        ws.cell(row=row_idx, column=8, value=record.get('closes_at', ''))
        ws.cell(row=row_idx, column=9, value=record.get('total_fund', ''))
        ws.cell(row=row_idx, column=10, value=record.get('project_size', ''))
        ws.cell(row=row_idx, column=11, value=record.get('project_funding_min', ''))
        ws.cell(row=row_idx, column=12, value=record.get('project_funding_max', ''))
        ws.cell(row=row_idx, column=13, value=record.get('expected_winners', ''))
        ws.cell(row=row_idx, column=14, value=record.get('funding_rules', ''))
        ws.cell(row=row_idx, column=15, value=record.get('sections_count', 0))
        ws.cell(row=row_idx, column=16, value=record.get('resources_count', 0))

        # Wrap text for description
        ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True, vertical='top')

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save
    wb.save(filename)
    logger.info(f"Excel file saved: {filename}")


def main():
    parser = argparse.ArgumentParser(description='Scrape Innovate UK competitions to Excel')
    parser.add_argument('--limit', type=int, default=None,
                        help='Maximum number of competitions to scrape')
    parser.add_argument('--output', type=str, default=None,
                        help='Output Excel filename (default: auto-generated)')
    args = parser.parse_args()

    print("=" * 60)
    print("INNOVATE UK SCRAPER - Excel Export")
    print("=" * 60)
    print()

    # Load URLs from file
    urls = load_urls_from_file()

    if not urls:
        logger.error("No competition URLs found. Add URLs to innovate_uk_urls.txt")
        sys.exit(1)

    logger.info(f"Loaded {len(urls)} URLs from innovate_uk_urls.txt")

    # Scrape competitions
    logger.info(f"Scraping {args.limit or len(urls)} competitions...")
    results = scrape_competitions(urls, limit=args.limit)

    if not results:
        logger.error("No competitions scraped successfully!")
        sys.exit(1)

    # Generate output filename
    if args.output:
        filename = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"innovate_uk_grants_{timestamp}.xlsx"

    # Export to Excel
    export_to_excel(results, filename)

    # Summary
    print()
    print("=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Competitions scraped: {len(results)}")
    print(f"Output file: {filename}")
    print()


if __name__ == "__main__":
    main()
